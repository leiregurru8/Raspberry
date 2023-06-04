
from flask import Flask, render_template
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

app = Flask(__name__)

@app.route('/')
def index():
    # Cargar el archivo Excel
    workbook = load_workbook('datos.xlsx')
    sheet = workbook.active

    # Obtener los datos de la hoja de cálculo
    data = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        humidity, temperature = row
        data.append({'humidity': humidity, 'temperature': temperature})

    # Obtener datos del tiempo: Scrapping

    #SOLICITUD HTP
    url= "https://www.timeanddate.com/weather/"
    response = requests.get(url)

    #ANALIZAR EL CONTENIDO DE LA RESPUESTA
    soup = BeautifulSoup(response.content, "html.parser")

    #ENCONTRAR EL CONTENIDO DESEADO MEDIANTE BS4
    temperatura = soup.find("span", class_ = "my-city__wtdesc").text.strip()  #los argumentos que cogemos

    

        # Renderizar el template HTML y pasar los datos a la plantilla
    return render_template('index.html', data=data, temperatura=temperatura)

if __name__ == '__main__':
    app.run(debug = True)
